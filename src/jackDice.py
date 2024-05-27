from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys

from selenium.webdriver import ActionChains
from selenium.webdriver.support.select import Select

import time
from time import sleep

import math
import random
import openpyxl
import config
import datetime

print("Welcome to JackDice Bot v2024.05.27")
options = webdriver.ChromeOptions()
options.add_experimental_option("detach", True)
options.add_argument("ignore-certificate-errors")
options.add_argument("--window-size=1920,1080")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--no-sandbox")
options.add_argument("--headless=new")

driver = webdriver.Chrome(options=options)
driver.implicitly_wait(10)

driver.execute_script("document.body.style.zoom = '0.5'")

username = config.username
password = config.password

def delay():
    time.sleep(random.randint(5,20))

driver.get("https://google.com")
delay()
driver.get("https://jacksclub.io")
print("Open Jacksclub.io")
delay()

# xPath Login
loginButton ='//*[@id="wrap"]/header/div/div/div[1]'
driver.find_element(By.XPATH, loginButton).click()
delay()

username_field = driver.find_element(By.NAME, value="username")
password_field = driver.find_element(By.NAME, value="password")

username_field.send_keys(username)
password_field.send_keys(password)

login_button = driver.find_element(By.XPATH, '//*[@id="modal-container"]/div[2]/div/div[1]/div[2]/div/div[2]/div[2]/div/form/div/div[4]')
login_button.click()
print("Login Jacksclub.io for " + username)
delay()

def pilihKoin():
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="wrap"]/main/article/div[1]/div'))).click()

    BTC = driver.find_element(By.XPATH, '//*[@id="floating"]/div/div/div/div[2]/div/div[1]/ul/li[1]/div/span[1]').text
    BCH = driver.find_element(By.XPATH, '//*[@id="floating"]/div/div/div/div[2]/div/div[1]/ul/li[2]/div/span[1]').text
    LTC = driver.find_element(By.XPATH, '//*[@id="floating"]/div/div/div/div[2]/div/div[1]/ul/li[3]/div/span[1]').text
    DOGE = driver.find_element(By.XPATH, '//*[@id="floating"]/div/div/div/div[2]/div/div[1]/ul/li[4]/div/span[1]').text
    ETH = driver.find_element(By.XPATH, '//*[@id="floating"]/div/div/div/div[2]/div/div[1]/ul/li[5]/div/span[1]').text
    XRP = driver.find_element(By.XPATH, '//*[@id="floating"]/div/div/div/div[2]/div/div[1]/ul/li[6]/div/span[1]').text
    TRX = driver.find_element(By.XPATH, '//*[@id="floating"]/div/div/div/div[2]/div/div[1]/ul/li[7]/div/span[1]').text
    USDT = driver.find_element(By.XPATH, '//*[@id="floating"]/div/div/div/div[2]/div/div[1]/ul/li[8]/div/span[1]').text
    BNB = driver.find_element(By.XPATH, '//*[@id="floating"]/div/div/div/div[2]/div/div[1]/ul/li[9]/div/span[1]').text
    JCC = driver.find_element(By.XPATH, '//*[@id="floating"]/div/div/div/div[2]/div/div[1]/ul/li[10]/div/span[1]').text

    delay()

    cekKoins = dict()
    if BTC != "":
        cekKoins['1'] = float(BTC)
    if BCH != "":
        cekKoins['2'] = float(BCH)
    if LTC != "":
        cekKoins['3'] = float(LTC)
    if DOGE != "":
        cekKoins['4'] = float(DOGE)
    if ETH != "":
        cekKoins['5'] = float(ETH)
    if XRP != "":
        cekKoins['6'] = float(XRP)
    if TRX != "":
        cekKoins['7'] = float(TRX)
    if USDT != "":
        cekKoins['8'] = float(USDT)
    if BNB != "":
        cekKoins['9'] = float(BNB)
    if JCC != "":
        cekKoins['10'] = float(JCC)

    cssCode = dict()
    cssCode['BTC'] = 1
    cssCode['BCH'] = 2
    cssCode['LTC'] = 3
    cssCode['DOGE'] = 4
    cssCode['ETH'] = 5
    cssCode['XRP'] = 6
    cssCode['TRX'] = 7
    cssCode['USDT'] = 8
    cssCode['BNB'] = 9
    cssCode['JCC'] = 10

    coinBalance = 0
    gudangKoin = config.coin.split(',')
    while float(coinBalance) <= 0:
        cekKoin = random.choice(gudangKoin)
        cekKoin = cekKoin.strip()
        cekKoinRandom = str(cssCode[cekKoin])
        delay()
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="floating"]/div/div/div/div[2]/div/div[1]/ul/li[' + cekKoinRandom + ']/div/span[2]'))).click()
        coinBalance = driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[1]/div/div[1]/div/div/span/span/span[1]').text

    print("Selected coin = " + cekKoin)
    
    return cekKoin

print("Open Dice Game Jacksclub.io")
delay()
driver.get("https://jacksclub.io/games/dice")
delay()

# sound Effect & instant Bet effect
if config.soundEffect == False or config.ambientMusic == False or config.instantBet == True:
    WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/ul/li[1]'))).click()

    if config.soundEffect == False:
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="floating"]/div/div/div/div[2]/div/div[1]/ul/li[1]'))).click()

    if config.ambientMusic == False:
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="floating"]/div/div/div/div[2]/div/div[1]/ul/li[2]'))).click()

    if config.instantBet == True:
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="floating"]/div/div/div/div[2]/div/div[1]/ul/li[3]'))).click()

# klik autobot
WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/ul/li[2]'))).click()

# lihat stat
WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/ul/li[3]'))).click()

action = ActionChains(driver)
statSource = driver.find_element(By.XPATH,'/html/body/div/div[2]/div[2]')
statTarget = driver.find_element(By.XPATH,'//*[@id="page"]/aside')
action.drag_and_drop(statSource, statTarget).perform()

def cekDice():
    if driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/footer/div[3]/div[1]/div/div[2]/input').is_enabled():
        return 1
    else:
        return 0

# Set auto dice
def setDice(bb: float = 0, multiplier: float = 1.0111, roll_over: int = 0, number_of_bets: int = 0, stop_on_profit: float = 0, stop_on_loss: float = 0, max_bet_amount: float = 0, stop_on_wins: int = 0, stop_on_losses: int = 0, on_win: int = 0, on_win_percent: float = 0, on_loss: int = 0, on_loss_percent: float = 0):
    coinBalance = driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[1]/div/div[1]/div/div/span/span/span[1]').text
    print("Current Balance = " + str(coinBalance) + " " + cekKoin)
    if(bb > 0 and multiplier > 0 and float(coinBalance) > 0):        
        if(multiplier <=1):
            multiplier = 1.0111

        multiplier = math.ceil(multiplier * 100) / 100

        driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/footer/div[2]/div[1]/input').clear()
        driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/footer/div[2]/div[1]/input').send_keys(multiplier)

        if float(config.bbMax) > 0:
            bbMax = float(config.bbMax)
            if bb > bbMax and bbMax > 0:
                bb = bbMax

        driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/footer/div[3]/div[1]/div/div[2]/input').clear()
        driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/footer/div[3]/div[1]/div/div[2]/input').send_keys(bb)

        roText = ""
        if(roll_over == 1):
            WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/footer/div[2]/div[2]'))).click()
            roText = "Set Random Roll Over/Under = Active"

        driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[1]/div[1]/input').clear()
        if(number_of_bets > 0):
            driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[1]/div[1]/input').send_keys(number_of_bets)
            number_of_bets = int(number_of_bets)
            nobText = "Set NOB = " + str(number_of_bets)

        driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[1]/div[2]/input').clear()
        if stop_on_profit > 0:
            driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[1]/div[2]/input').send_keys(stop_on_profit)

        driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[1]/div[3]/input').clear()
        if stop_on_loss > 0:
            driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[1]/div[3]/input').send_keys(stop_on_loss)

        driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[1]/div[4]/input').clear()
        if max_bet_amount > 0:
            driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[1]/div[4]/input').send_keys(max_bet_amount)

        driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[1]/div[5]/input').clear()
        if stop_on_wins > 0:
            driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[1]/div[5]/input').send_keys(stop_on_wins)

        driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[1]/div[6]/input').clear()
        if stop_on_losses > 0:
            driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[1]/div[6]/input').send_keys(stop_on_losses)

        while driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[1]/input').is_enabled():
            WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[1]/div[2]/button/div[1]'))).click()

        if on_win > 0:
            if(on_win == 1): # decrease
                WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[1]/div[2]/button/div[1]'))).click()
                
                on_win_percent = round(on_win_percent * 100) / 100
                driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[1]/input').clear()
                driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[1]/input').send_keys(on_win_percent)

            elif on_win == 2: # increase
                WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[1]/div[2]/button/div[1]'))).click()
                WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[1]/div[2]/button/div[1]'))).click()

                on_win_percent = math.ceil(on_win_percent * 100) / 100
                driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[1]/input').clear()
                driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[1]/input').send_keys(on_win_percent)
                
        while driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[2]/input').is_enabled():
            WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[2]/div[2]/button/div[1]'))).click()

        if on_loss > 0:
            if(on_loss == 1): # decrease
                WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[2]/div[2]/button/div[1]'))).click()

                on_loss_percent = round(on_loss_percent * 100) / 100
                driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[2]/input').clear()
                driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[2]/input').send_keys(on_loss_percent)
                
            elif on_loss == 2: # increase
                on_loss_percent = math.ceil(on_loss_percent * 100) / 100
                WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[2]/div[2]/button/div[1]'))).click()
                WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[2]/div[2]/button/div[1]'))).click()

                driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[2]/input').clear()
                driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[2]/input').send_keys(on_loss_percent)

        driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/footer/div[2]/div[1]/input').send_keys(Keys.TAB)
        driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/footer/div[3]/div[1]/div/div[2]/input').send_keys(Keys.TAB)
        delay()
        poText = driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/footer/div[2]/div[1]/input').get_attribute("value")
        bbText = driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/footer/div[3]/div[1]/div/div[2]/input').get_attribute("value")
        bbText = float(bbText)

        print("Set PO = x" + str(poText))
        print("Set BB = " + f'{bbText:.8f}' + " " + cekKoin)

        if roText.strip() != "":
            print(roText)

        if(number_of_bets > 0):
            print(nobText)

        if stop_on_profit > 0:
            sopText = driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[1]/div[2]/input').get_attribute("value")
            print("Set Stop Profit = " + str(sopText))

        if stop_on_loss > 0:
            solText = driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[1]/div[3]/input').get_attribute("value")
            print("Set Stop Loss = " + str(solText))

        if max_bet_amount > 0:
            mbaText = driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[1]/div[4]/input').get_attribute("value")
            print("Set Max Bet = " + str(mbaText))

        if stop_on_wins > 0:
            sowText = driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[1]/div[5]/input').get_attribute("value")
            print("Set Stop on Win = " + str(sowText))

        if stop_on_losses > 0:
            solsText = driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[1]/div[6]/input').get_attribute("value")
            print("Set Stop on Losses = " + str(solsText))

        if driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[1]/input').is_enabled():
            owpText = driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[1]/input').get_attribute("value")
            if on_win == 1:
                owpText = "On Win dec = " + str(owpText) + "%"
            else:
                owpText = "On Win inc = " + str(owpText) + "%"

            print(owpText)


        if driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[2]/input').is_enabled():
            olpText = driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[2]/input').get_attribute("value")
            if on_loss == 1:
                olpText = "On Loss dec = " + str(olpText) + "%"
            else:
                olpText = "On Loss inc = " + str(olpText) + "%"
                
            print(olpText)

        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/footer/div[3]/div[2]'))).click()

    else:
        return 0

def readStat():
    bacaStat = dict()
    bacaStat['profit'] = driver.find_element(By.XPATH, '//*[@id="app"]/div[2]/div/div/ul/li[1]/span/span/span').text
    bacaStat['wager'] = driver.find_element(By.XPATH, '//*[@id="app"]/div[2]/div/div/ul/li[2]/span/span/span').text
    bacaStat['bets'] = driver.find_element(By.XPATH, '//*[@id="app"]/div[2]/div/div/ul/li[3]/span').text
    bacaStat['win'] = driver.find_element(By.XPATH, '//*[@id="app"]/div[2]/div/div/ul/li[4]/span').text
    bacaStat['loss'] = driver.find_element(By.XPATH, '//*[@id="app"]/div[2]/div/div/ul/li[5]/span').text

    coinBalance = driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[1]/div/div[1]/div/div/span/span/span[1]').text
    bbText = driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/footer/div[3]/div[1]/div/div[2]/input').get_attribute("value")
    poText = driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/footer/div[2]/div[1]/input').get_attribute("value")

    olpText = "0"
    if driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[2]/input').is_enabled():
        olpText = driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[2]/input').get_attribute("value")

    owpText = "0"
    if driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[1]/input').is_enabled():
        owpText = driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/div/div[2]/div[1]/input').get_attribute("value")

    tglUTC = datetime.datetime.now(datetime.timezone.utc).strftime('%Y-%m-%d')
    tglJamUTC = datetime.datetime.now(datetime.timezone.utc).strftime('%Y-%m-%d %H:%M:%S')
    fileDataProfit = config.fileDataProfit + cekKoin + "-" + tglUTC + ".xlsx"
    try:
        wb = openpyxl.load_workbook(fileDataProfit)
        sheet = wb["Sheet"]
        dataSheet = sheet.max_row + 1
        totalProfit = sheet.cell(sheet.max_row, sheet.max_column).value

    except FileNotFoundError:
        wb = openpyxl.Workbook() 
        sheet = wb.active 

        sheet['A1'].value = "No"        
        sheet['B1'].value = "Date Time" 
        sheet['C1'].value = "Balance" 
        sheet['D1'].value = "BB" 
        sheet['E1'].value = "PO" 
        sheet['F1'].value = "OWP"
        sheet['G1'].value = "OLP"
        sheet['H1'].value = "Bets"
        sheet['I1'].value = "Wins"
        sheet['J1'].value = "Losses"
        sheet['K1'].value = "Wagered"
        sheet['L1'].value = "Profit"
        sheet['M1'].value = "Total Profit"

        totalProfit = 0
        dataSheet = 2
        sheet.max_row = 0

    if float(bacaStat['profit']) > 0:
        totalProfit = float(bacaStat['profit']) + float(totalProfit)

    sheet['A' + str(dataSheet)].value = "1"        
    sheet['B' + str(dataSheet)].value = tglJamUTC 
    sheet['C' + str(dataSheet)].value = coinBalance 
    sheet['D' + str(dataSheet)].value = bbText
    sheet['E' + str(dataSheet)].value = poText
    sheet['F' + str(dataSheet)].value = owpText
    sheet['G' + str(dataSheet)].value = olpText
    sheet['H' + str(dataSheet)].value = bacaStat['bets']
    sheet['I' + str(dataSheet)].value = bacaStat['win']
    sheet['J' + str(dataSheet)].value = bacaStat['loss']
    sheet['K' + str(dataSheet)].value = bacaStat['wager']
    sheet['L' + str(dataSheet)].value = bacaStat['profit']
    sheet['M' + str(dataSheet)].value = totalProfit

    wb.save(fileDataProfit) 
    wb.close()

    if float(bacaStat['profit']) >= 0:
        textProfit = "Profit"
        textProfitX = "Dice failed"
        if float(bacaStat['profit']) > 0:
            textProfitX = "Dice success with new profits"
    else:
        textProfit = "Loss"
        textProfitX = "Dice success but loss & will be recovered"

    
    if float(bacaStat['profit']) < 0 and sheet.max_row > 0:
        # print("max row = " + str(sheet.max_row))
        # print("baca = " + bacaStat['win'])
        # if (int(sheet.cell["I" + str(sheet.max_row)].value) - int(bacaStat['win'])) > 1:
        totalProfit = totalProfit + float(bacaStat['profit'])
            # bacaStat['profit'] = float(sheet.cell["L" + str(sheet.max_row)].value) - bacaStat['profit']
            # bacaStat['bets'] = float(sheet.cell["H" + str(sheet.max_row)].value) - int(bacaStat['bets'])
            # bacaStat['wager'] = float(sheet.cell["K" + str(sheet.max_row)].value) - int(bacaStat['wager'])
            # bacaStat['win'] = int(sheet.cell["I" + str(sheet.max_row)].value) - int(bacaStat['win'])
            # bacaStat['loss'] = int(sheet.cell["J" + str(sheet.max_row)].value) - int(bacaStat['loss'])

    bacaStat['profit'] = float(bacaStat['profit'])
    print(textProfit + " = " + f'{bacaStat['profit']:.8f}' + " " + cekKoin)
    print("Wagered = " + bacaStat['wager'] + " " + cekKoin)
    print("Bets = " + bacaStat['bets'] + " rolls")
    print("Win = " + bacaStat['win'] + "x")
    print("Losses = " + bacaStat['loss'] + " rolls")
    print("Total profit = " + f'{totalProfit:.8f}')
    print(textProfitX)

    if config.showfileProfit == True:
        print("Write data to file " + fileDataProfit)

    print("Data is saved successfully") 
    print("======================================")
    print("")

    return bacaStat

def resetSeed():
    # reset seed
    cekRoll = driver.find_element(By.XPATH, '//*[@id="app"]/div[2]/div/div/ul/li[3]/span').text
    if int(cekRoll) > int(config.seedReset) and random.randint(1, 10) == 1:
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/ul/li[5]'))).click()
        delay()
        if int(driver.find_element(By.XPATH, '//*[@id="modal-container"]/div[2]/div/div[1]/div[2]/div/div[2]/div[1]/div[2]/div/input').get_attribute("value")) >= int(config.seedReset):
            WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="modal-container"]/div[2]/div/div[1]/div[2]/div/div[3]/div/div[1]/div/div'))).click()
            print("Reset seed success")
            delay()

        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="modal-container"]/div[2]/div/div[1]/div[2]/div/div[1]'))).click()

def checkPO(lastProfit: float = 0, lastPO: float = 0):
    tabDataPOs = config.tabDataPO.split(',')
    profitFilters = config.profitFilter.split(',')
    
    tabPO = ""
    tabPOs = dict()
    i = 0
    for pF in profitFilters:
        tabPOs[pF] = tabDataPOs[i]
        if float(lastProfit) < float(pF) and str(tabPO) == "":
            tabPO = tabDataPOs[i]
        i = i + 1

    if tabPO == tabDataPOs[i - 1]:
        del tabDataPOs[0]
        tabPO = random.choice(tabDataPOs)

    # reset stat
    if float(lastProfit) > 0:
        WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="app"]/div[2]/div/div/header/nav/div[1]'))).click()
        print("Reset stat success")

    bbText = driver.find_element(By.XPATH, '//*[@id="wrap"]/main/article/div[2]/div[1]/div/div/div/div[1]/div/footer/div[3]/div[1]/div/div[2]/input').get_attribute("value")
    if float(bbText) > 0:
        bb = round(float(bbText))
    else:
        bb = 0.00000001
    
    if lastProfit >= 0 and float(config.bbMin) > 0 and bb < float(config.bbMin):
        bb = float(config.bbMin)

    if lastProfit >= 0 and bb > float(config.bbMax) and float(config.bbMax) > 0:
        bb = float(config.bbMax)

    workbook = openpyxl.load_workbook(config.fileDataPO)
    sheet = workbook[tabPO.strip()]

    lastPO = sheet.max_row
    print("Choose PO formula = " + tabPO)

    dataDice = random.randint(2, lastPO)
    if (int(sheet.cell(dataDice, 13).value) == 2 or int(sheet.cell(dataDice, 13).value) == 1) and float(sheet.cell(dataDice, 14).value) > 0:
        bbMin = math.ceil((100 / float(sheet.cell(dataDice, 14).value))) / 100000000
        if bb > float(config.bbMax) and float(config.bbMax) > 0:
            while float(config.bbMin) < bbMin and float(config.bbMin) > 0:
                # tabPO = tabDataPOs[random.randint(1, len(tabDataPOs) - 1)]
                # sheet = workbook[tabPO.strip()]
                # lastPO = dataDice

                if (int(sheet.cell(dataDice, 13).value) == 2 or int(sheet.cell(dataDice, 13).value) == 1) and float(sheet.cell(dataDice, 14).value) > 0:
                    bbMin = math.ceil((100 / float(sheet.cell(dataDice, 14).value))) / 100000000
                else:
                    if float(config.bbMin) > 0:
                        bbMin = config.bbMin
                    else:
                        bbMin = 0.00000001
            
                if float(config.bbMin) > 0 and float(config.bbMin) > bbMin:
                    bbMin = float(config.bbMin)
                    
                if float(config.bbMax) > 0 and float(config.bbMax) < bbMin:
                    bbMin = float(config.bbMax)

        if lastProfit >= 0:
            bbIncChoice = 0.00
            bbIncs = config.bbInc.split(',')

            if float(bbIncs[0]) >= 1 and float(bbIncs[1]) >= 1:
                bbIncChoice = random.uniform(float(bbIncs[0]),float(bbIncs[1]))
            else:
                bbIncChoice = float(bbIncs[0])

            if float(bbIncChoice) > 1:
                bb = bb * float(bbIncChoice)
        else:
            bbIncChoice = 0
            bbIncs = config.bbIncLoss.split(',')
            
            if float(bbIncs[0]) >= 1 and float(bbIncs[1]) >= 1:
                bbIncChoice = random.uniform(float(bbIncs[0]),float(bbIncs[1]))
            else:
                bbIncChoice = float(bbIncs[0])
            
            if float(bbIncChoice) > 1:
                bb = bb * float(bbIncChoice)

        if bb < bbMin:
            bb = bbMin

        if bb > float(config.bbMax) and float(config.bbMax) > 0:
            bb = float(config.bbMax)

    dataDices = dict()
    dataDices['dataDice'] = dataDice
    dataDices['bb'] = bb
    dataDices['multiplier'] = float(sheet.cell(dataDice, 3).value)
    dataDices['roll_over'] = int(sheet.cell(dataDice, 4).value)
    dataDices['number_of_bets'] = int(sheet.cell(dataDice, 5).value)
    dataDices['stop_on_profit'] = float(sheet.cell(dataDice, 6).value)
    dataDices['stop_on_loss'] = float(sheet.cell(dataDice, 7).value)
    dataDices['max_bet_amount'] = float(sheet.cell(dataDice, 8).value)
    dataDices['stop_on_wins'] = int(sheet.cell(dataDice, 9).value)
    dataDices['stop_on_losses'] = int(sheet.cell(dataDice, 10).value)
    dataDices['on_win'] = int(sheet.cell(dataDice, 11).value)
    dataDices['on_win_percent'] = float(sheet.cell(dataDice, 12).value)
    dataDices['on_loss'] = int(sheet.cell(dataDice, 13).value)
    dataDices['on_loss_percent'] = float(sheet.cell(dataDice, 14).value)
    delay()

    return dataDices

profitTotal = 0
lastProfit = 0
lastPO = 0

stopProfit = float(config.stopProfit)
stopLoss = float(config.stopLoss)

while (profitTotal < stopProfit and stopProfit != 0) or (profitTotal > stopLoss and stopLoss != 0) or stopProfit == 0 or stopLoss == 0 :
    if lastProfit >= 0:
        cekKoin = pilihKoin()
        profitTotal = 0
        lastPO = 0
        dataDices = checkPO(lastProfit, lastPO)

    elif lastProfit < 0:
        dataDices = checkPO(lastProfit, lastPO)

    setDice(dataDices['bb'], dataDices['multiplier'], dataDices['roll_over'], dataDices['number_of_bets'], dataDices['stop_on_profit'], dataDices['stop_on_loss'], dataDices['max_bet_amount'], dataDices['stop_on_wins'], dataDices['stop_on_losses'], dataDices['on_win'], dataDices['on_win_percent'], dataDices['on_loss'], dataDices['on_loss_percent'])
    
    while cekDice() != 1:
        print(datetime.datetime.now(datetime.timezone.utc).strftime('%Y-%m-%d %H:%M:%S') + " Jack Dice running...")
        resetSeed()
        delay()

    bacaStat = readStat()
    lastProfit = float(bacaStat['profit'])

else:
    print("Dice End. Total profit = " + str(profitTotal))
